##初始設定
import numpy as np
SLE_domain = np.array([['Constitutional'], ['Hematologic'], ['Neuropsychiatric'], ['Mucocutaneous'], ['Serosal'], ['Musculoskeletal'], ['Renal'], ['Antiphospholipid antibodies'], ['Complement_proteins'], ['SLE-specific antibodies']])
# condition nparray(10*4 = 7*4 + 3*4)
SLE_ccond = np.array([['Fever', None, None, None], ['Leukopenia', 'Thrombocytopenia', 'Autoimmune hemolysis', None], ['Delirium', 'Psychosis', 'Seizure', None], ['Non-scarring alopecia', 'Oral ulcers', 'Subacute cutaneous OR discoid lupus', 'Acute cutaneous lupus'], ['Pleural or pericardial effusion', 'Acute pericarditis', None, None], ['Joint involvement',None, None, None], ['Proteinuria = ___ g/24hours by 24-hour urine?\n*Please input a positive integer or float.', 'Which class best subscribes your status of lupus nephritis according to renal biopsy?\n*Please input a integer from 0 to 6("0":Normal ; "1"-"6":Class 1-6 lupus nephritis)', None, None]])
SLE_icond = np.array([['Anti-cardiolipin antibodies OR Anti-β2GP1 antibodes OR Lupus anticoagulant', None, None, None], ['Low C3 OR low C4', 'Low C3 AND low C4', None, None],['Anti-dsDNA antibody OR Anti-Smith antibody', None, None, None]])
SLE_cond = np.vstack((SLE_ccond, SLE_icond))
# weight array
SLE_weight = np.array([[2, 0, 0, 0], [3, 4, 4, 0], [2, 3, 5, 0], [2, 2, 4, 6], [5, 6, 0, 0], [6, 0, 0, 0], [4, 8, 10, 0], [2, 0, 0, 0], [3, 4, 0, 0], [6, 0, 0, 0]])
# Ans array
SLE_ans = np.array(None, (object, [10, 4]))
# 製作問答loop_control_array(元素為各domain之condition個數) >>用於SLE各condition問答  
SLE_loop_control = np.count_nonzero(SLE_cond, axis=1)

### 各流程建立function
## 在初始介面先使用二分法確認使用者狀態(類似分類樹概念):
def Qstatus():
    global name, sickid
    name = input('What\'s your name?')
    sickid = input('What\'s your medical record number?') #詢問病歷號  
    print('Have you been diagnosed as SLE?')
    sta = 'str'
    while sta != 'Y' and sta != 'N':
        sta = input('Please input "Y" for YES and "N" for NO.(case-insensitive大小寫不拘) ').strip()
        sta = sta.capitalize() 
    return sta

status3 = 'N'
def QbeforeDAI():
    global status3
    sta1 = 'str'
    print('Would you like to further understand your SLE disease severity?')
    while sta1 != 'Y' and sta1 != 'N':
        sta1 = input('Please input "Y" for YES and "N" for NO.(case-insensitive大小寫不拘) ').strip()
        sta1 = sta1.capitalize() 
    status3 = sta1       
    return status3 
##entry_criterion_防呆裝置  
titer = 0
def entry_criterion():
    global titer
    print('''\nWelcome to SLE diagnosis assistant! We will predict whether or not you are classified as SLE before long.
 *Instructions for users: This App serves only as a diagnosis assistant, which is not designed for diagnosis or treatment decisions. Diagnosis of SLE remains the purview of an appropriately trained physician evaluating an individual patient.''')
    print('\nHere comes entry criteron.')
    while type(titer) != type(6.6) or titer <= 0:
        try:
            titer = float(input('ANA at a titer of 1:"?" on HEp-2 cells.(Please input a number of type"float".)').strip())
            if titer <= 0:
                print('You should input a positive number!')
            continue
        except Exception as ex:   # Exception為各種exceptions的superclass；將此exception存在ex變數中!!!   
            print(type(ex))          #當不確定例外種類，且想知道例外的原因時，可利用左邊範例的方法。
            print('Error!', ex)    
    return titer    

## additive_criteria問答 (利用巢狀while loop 處理與使用者的問答以及答題紀錄，適用於clinical & immunological)     
# 以SLE_domain、SLE_cond、SLE_ccond、SLE_ans、SLE_loop_control作為argument，考慮與SLEDAI_condition_QA合併(?)
def ConditionCriteria_QA(domain_array, CondArray, CCondArray, ans_array, LoopControlArray): 
    global SLE_ans
    print('\nAdditive criteria are needed before final diagnosis.')
    print(f'''There will be a survey about conditions in clinical or immunological domains.\n
*Here are something you should be informed of before the questionaire:
1. Please input correct key words for the App to run smoothly!
 "Y" stands for "Yes"(i.e. a patient meets the condition); "N" for "No"(i.e. one doesn't meet the criteria); "NI" for "no information" 
 "B" or "F" respectively for "jump *Back or *Forward to another question"
2. case-insensitive大小寫不拘    
3. All criteria are only to be counted if SLE is thought to be the most likely cause of the manifestation； occurrence of a criterion on at least one occasion is sufficient.
~ Thank you for your cooperation! ~''')   
    i = 0
    while i < len(domain_array):
        if i == 0:
            print(f'\n*Clinacal Conditions Questionaire:')
        if i == len(CCondArray): 
            print(f'\n*Immunological Conditions Questionaire:')
        print(f'\n{(i%len(CCondArray))+1}. Questions about {domain_array[i, 0]} domain:') #字串格式化
        j = 0
        while j < LoopControlArray[i]:     # while j < len(CondArray[i]) and CondArray[i, j] is not None: 也可以
            if i == 6:   
                print('In this special domain, you should read the instruction carefully and fill in the correct "Number"!')  
                print(f'({(i%len(CCondArray))+1}-{j+1}) {CondArray[i, j]} (Current Ans: "{ans_array[i, j]}")')      
                check = input(f'(specific Number/ NI/ B/ F)  ').strip().upper()
                if check == "NI":
                    ans_array[i, j] = check   ##
                    j = j + 1  
                elif check == "B": 
                    if j == 0:                   #當輸入B或F時，讓使用者知道自己正跳至別題檢查，並可更正答案(v)(*希望呈現出目前答案給使用者參考?)
                        if i != 0:
                            i -= 1 
                            j = LoopControlArray[i] - 1 
                        else:
                            print("*This is the first question!")            
                    else:
                        j -= 1
                elif check == "F":
                    if j == LoopControlArray[i] - 1:
                        if i != len(domain_array) - 1:
                            i += 1
                            j = 0
                        else:
                            print("*This is the last question!")    
                    else:
                        j += 1
                else:
                    # if not isinstance(check, float) or titer <= 0:
                    try:
                        check = float(check)
                        if check < 0:
                            print('You shouldn\'t input a negative number!')
                            j = j
                        else:
                            ans_array[i, j] = check
                            j += 1  
                    except ValueError:
                        print("Wrong key word!")
                    j = j 
            else:
                print(f'({(i%len(CCondArray))+1}-{j+1}) {CondArray[i, j]} ? (Current Ans: "{ans_array[i, j]}")')      #是否附註該condition的定義給使用者作確認(?)/back&forward(V)/預設no information(v)
                check = input(f'(Y/ N/ NI/ B/ F)  ').strip().upper()
                if check == "Y":
                    ans_array[i, j] = check   #先用Y、N組成ans_array，之後再process為ans_Number_array(0, 1)，再與SLEDAI_weight相乘，得到ans_Weight_array
                    j = j + 1
                elif check == "N":
                    ans_array[i, j] = check
                    j = j + 1
                elif check == "NI":
                    ans_array[i, j] = check   ##
                    j = j + 1  
                elif check == "B": 
                    if j == 0:                   #當輸入B或F時，讓使用者知道自己正跳至別題檢查，並可更正答案(v)(*希望呈現出目前答案給使用者參考?)
                        if i != 0:
                            i -= 1 
                            j = LoopControlArray[i] - 1 
                        else:
                            print("*This is the first question!")            
                    else:
                        j -= 1
                elif check == "F":
                    if j == LoopControlArray[i] - 1:
                        if i != len(domain_array) - 1:
                            i += 1
                            j = 0
                        else:
                            print("*This is the last question!")    
                    else:
                        j += 1
                else:
                    print("Wrong key word!")
                    j = j  
        #惟最後一題無法更改>考慮新增一項for最後檢查?        
        i += 1
        SLE_ans = ans_array
    return ans_array

##ans confirmation(分為clinical & immunological)
def Ans_Confirmation(CondArray, ans_array, LoopControlArray):
  # 1. 呈現出目前答題狀況(題號, condition, ans)
  print('\n* The following dataframe shows your current answers to each question, and we\'ll help you confirm them before the next step.')
  print('Q_Num  ', "Condition".ljust(85), 'Current_Ans\n', '-' * 125)
  for i in range(len(CondArray)):
    j = 0
    while j < LoopControlArray[i]:
      a = CondArray[i, j].find('?')
      if a != -1:
        print(f'{i+1}-{j+1:<6}{CondArray[i, j][0:a]:<85} {ans_array[i, j]}')
      else:
        print(f'{i+1}-{j+1:<6}{CondArray[i, j]:<85} {ans_array[i, j]}')
      j += 1
  # 2. 修正
  confirm2 = None
  while confirm2 != 'Y':
    confirm1 = None
    while confirm1 != 'Y' and confirm1 != 'N':
      confirm1 = input('Is there anything wrong that you\'d like to modify? (Please input "Y" or "N")').strip().upper()
    while confirm1 == 'Y':
      while True:
        QNum = input('Please input the question number in correct format! (eg. 1-1, 3-3, or 7-2 ; one at a time!)').strip()
        if ',' in QNum:
          qnlist = []
          for qn in QNum.split(','):
            qnlist.append(qn.strip())
        else:
          qnlist = [QNum]    
        try:
          for k in range(len(qnlist)):
            i, j = qnlist[k].split('-')
            i = int(i) - 1
            j = int(j) - 1
            qnlist[k] = (i, j) 
          if 0 <= i < len(CondArray) and 0 <= j < LoopControlArray[i]:
            break
          else:
            print('Wrong question number!')
            continue  
        except Exception as ex:   # Exception為各種exceptions的superclass；將此exception存在ex變數中!!!  
          print('Wrong word, please try again!') 
          print(type(ex), ':', ex)  
      for i, j in qnlist:          
        if i == 6:   
          print('In this special domain, you should read the instruction carefully and fill in the correct "Number"!')
          while True:
            print(f'({i+1}-{j+1}) {CondArray[i, j]} (Current Ans: "{ans_array[i, j]}")')      
            check = input(f'(specific Number/ NI)  ' ).strip().upper()
            if check == "NI":
                ans_array[i, j] = check   ##
                break
            else:
              try:
                check = float(check)
                if check < 0:
                  print('You shouldn\'t input a negative number!')
                else:
                  ans_array[i, j] = check
                  break
              except Exception as ex:   # Exception為各種exceptions的superclass；將此exception存在ex變數中!!!   
                print("You just filled in a wrong key word, please be more careful and try again!")
                print('Error_type:',type(ex))          
                print('Error!', ex)
        else:
          print(f'({i+1}-{j+1}) {CondArray[i, j]} ? (Current Ans: "{ans_array[i, j]}")')      #是否附註該condition的定義給使用者作確認(?)/back&forward(V)/預設no information(v)
          while True:
            check = input(f'(Y/ N/ NI)  ' ).strip().upper()
            if check == "Y":
                ans_array[i, j] = check   #先用Y、N組成ans_array，之後再process為ans_Number_array(0, 1)，再與SLEDAI_weight相乘，得到ans_Weight_array
                break
            elif check == "N":
                ans_array[i, j] = check
                break
            elif check == "NI":
                ans_array[i, j] = check   ##
                break
            else:
                print("You just filled in a wrong key word, please be more careful and try again!")
      confirm1 = input('Is there anything else wrong that you\'d like to modify? (Please input "Y" or "N")').strip().upper()
    #再次確認目前ans 
    print('Again! The following dataframe shows your current answers to each question. Please double check them!') 
    print('Q_Num  ', "Condition".ljust(85), 'Current_Ans\n', '-' * 125)
    for i in range(len(CondArray)):
      j = 0
      while j < LoopControlArray[i]:
        a = CondArray[i, j].find('?')
        if a != -1:
          print(f'{i+1}-{j+1:<6}{CondArray[i, j][0:a]:<85} {ans_array[i, j]}')
        else:
          print(f'{i+1}-{j+1:<6}{CondArray[i, j]:<85} {ans_array[i, j]}')
        j += 1
    #最後確認
    confirm2 = None
    while confirm2 != 'Y' and confirm2 != 'N':
      confirm2 = input('Have you already confirmed that all answers correspond with your current health status? (Please input "Y" or "N")').strip().upper()
  print('Answer confirmation has finished! Please wait a moment for the results!')  
#---------------------------------------------------------------------------------------------------------------#
## 處理各domains的weights(適用於clinical & immunological)
## key idea: 先用Y、N組成ans_array(v)，之後再process為ans_Number_array(-1無此問題, 0無症狀或無資料, 1有症狀)，再與SLEDAI_weight相乘，得到ans_Weight_array
#process為ans_Number_array(-1, 0, 1)
def Change_YN_to_Number(Ans_Array):
  global SLE_ansNum, SLE_ansNI, SLE_weight, SLE_cond, SLE_ans_Weight
  SLE_ansNum = SLE_weight.copy()
  SLE_ansNI = []
  for i in range(len(SLE_cond)):
    SLE_ansNI.append([])    # 此nested_list(SLE_ansNI) 將用來放置各domain無資料的index

  for i in range(len(Ans_Array)):
    for j in range(len(Ans_Array[i])):
      if i == 6:  # renal domain就你最特別
        if j == 0:  #proteinuria
          if isinstance(Ans_Array[i, j], float):  #可能為float的只可能是proteinuria(0)、lupus nephritis(1)
            SLE_ansNum[i, j] = 1 if Ans_Array[i, j] > 0.5 else 0
          else:  #只可能是NI
            SLE_ansNum[i, j] = 0
            SLE_ansNI[i].append(j)
        elif j == 1:  #lupus nephritis (2,5>>8 pts；3,4>>10 pts)，好複雜QQ
          if isinstance(Ans_Array[i, j], float): 
            if Ans_Array[i, j] == 2 or Ans_Array[i, j] == 5:
              SLE_ansNum[i, j] = 1
              SLE_ansNum[i, j+1] = 0
            elif Ans_Array[i, j] == 3 or Ans_Array[i, j] == 4:
              SLE_ansNum[i, j] = 0
              SLE_ansNum[i, j+1] = 1
            else:
              SLE_ansNum[i, j] = 0
              SLE_ansNum[i, j+1] = 0  
          else:   #只可能是NI
            SLE_ansNum[i, j] = 0
            SLE_ansNum[i, j+1] = 0
            SLE_ansNI[i].extend([j, j+1])    
        else:  #沒題目None  
          SLE_ansNum[i, 3] = -1    #否則會出問題，[6, 2]會蓋過[6, 1]時產生的數字
      else:  
        if Ans_Array[i, j] == "Y":
          SLE_ansNum[i, j] = 1
        elif Ans_Array[i, j] == "N":
          SLE_ansNum[i, j] = 0
        elif Ans_Array[i, j] == "NI":
          SLE_ansNum[i, j] = 0
          SLE_ansNI[i].append(j)  
        else:
          SLE_ansNum[i, j] = -1 
  SLE_ans_Weight = np.multiply(SLE_ansNum, SLE_weight)              
  return SLE_ansNum, SLE_ansNI, SLE_ans_Weight
#找出各domain最高分
def Find_Domain_Max(AnsWeight):
  global Domain_weight, total_score
  Domain_weight = []
  for i in AnsWeight:
    Domain_weight.append(max(i))
  total_score = sum(Domain_weight)
  return Domain_weight, total_score  
# 找到weight最高的domain  >>判定main determinant instead of main cause!!! 
# 或highest_score = max(Domain_weight)  
def Find_Main_Detreminamt(DomainWeight, Domain, SLECCond):
  global highest_score, main_cdeterminant, main_ideterminant
  highest_score = max(DomainWeight)
  main_cdeterminant = []
  main_ideterminant = []
  for i in range(len(DomainWeight)):
    if DomainWeight[i] == highest_score:
      if i < len(SLECCond):
        main_cdeterminant.append(f'{Domain[i, 0]} domain')
      else:
        main_ideterminant.append(f'{Domain[i, 0]} domain')          
  return main_cdeterminant, main_ideterminant  

# SLE final diagnosis: SLE classification requires at least one clinical criterion and ≥10 points.
def SLE_Classification_Advice(total_score, DomainWeight, SLECCond):
  global main_cdeterminant, main_ideterminant, sleresult
  Nccri = 0  #確認是否完全沒有達到任何clinical_criteria
  for i in DomainWeight[0:len(SLECCond)]:
    if i == 0:
      Nccri += 1
  print(f'\n*Diagnostic result:')    
  if total_score >= 10:
    if Nccri == len(SLECCond):
      print(f'No. This patient is probably not classified as SLE because he/she doesn\'t meet any clinical criterion.')
      sleresult = 'No'
    else:
      #不知道使用全域變數是否需要傳入該引數
      print(f'''Yes. This patient is probably classified as SLE.
Main determinants: (1)clinical: {main_cdeterminant} ; (2)immunology: {main_ideterminant}  
Each gets {highest_score} pts of total {total_score} pts and accounts for {highest_score/total_score * 100:.2f} % of total weights.''') 
      sleresult = 'Yes'
  else:
    print(f'''No. This patient is probably not classified as SLE.
The total weights add up to only {total_score} pts, which doesn't reach the standard 10 pts for SLE diagnosis.''')
    sleresult = 'No'
  return sleresult  
# No_Information_Suggestion
def NI_Suggestion1(Ans_NI, DomainArray, CondArray):  # ConditionArray(SLE_cond) ; DomainArray(SLE_domain)
  print(f'\nThe conditions with no information, if any, are as follows:') 
  list1 = []
  for i in range(len(Ans_NI)):
    if Ans_NI[i] == []:
      continue
    else:
      for j in Ans_NI[i]:
        print(f'{CondArray[i, j]} (in {DomainArray[i]} domain)')
        list1.append(f'{CondArray[i, j]} (in {DomainArray[i]} domain)')
  print(f'\n*IF you have some questions with NO INFORMATION, those conditions are considered negative automatically by the App and thus it is likely that the patient\'s disease severity is underestimated. In this case, We suggest that you take further examinations to make up the dificiency and then you can get more accurate predictions from the App! Thanks!''')
  return list1 

#---------------------------------------------------------------------------------------------#
## B. SLEDAI:  #盡量不要重複問(Fever, Thrombocytopenia, Psychosis, Seizure, Proteinuria)
#如果使用者有意願 >>> 用SLEDAI進一步嚴重度分級 / 先不做科別方類
import numpy as np
SLEDAI_cond = np.array(["Seizure(recent onset)", "Psychosis", "Organic brain syndrome", "Visual disturbance", "Cranial nerve disorder", "Lupus headache", "CVA(new onset)", "Vasculitis", "Arthritis", "Myositis", "Urinary casts", "Hematuria", "Proteinuria", "Pyuria", "Rash", "Alopecia", "Mucosal ulcers", "Pleurisy", "Pericarditis", "Low complement", "Increased DNA binding", "Fever", "Thrombocytopenia", "Leukopenia"])
SLEDAI_def = np.array(["Recent onset, exclude metabolic, infections, or drug causes.", "Altered ability to function in normal activity due to severe disturbance in the perception of reality. Exclude uremia and drug causes.", " Altered mental function with impaired orientation, memory, or other intellectual function, with rapid onset and fluctuating clinical features, inability to sustain attention to environment, plus at least 2 of the following: perceptual disturbance, incoherent speech, insomnia or daytime drowsiness, or increased or decreased psychomotor activity. Exclude metabolic, infectious, or drug causes.", "Retinal changes of SLE. Exclude hypertension, infection, or drug causes", "New onset of sensory or motor neuropathy involving cranial nerves.", "Severe, persistent headache; may be migrainous, but must be nonresponsive to narcotic analgesia.", "New onset of cerebrovascular accident(s). Exclude arteriosclerosis.", "Ulceration, gangrene, tender finger nodules, periungual infarction, splinter hemorrhages, or biopsy or angiogram proof of vasculitis.", "≥2 joints with pain and signs of inflammation (i.e., tenderness, swelling, or effusion).", "Proximal muscle aching/weakness, associated with elevated creatine phosphokinase/aldolase or electromyogram changes or a biopsy showing myositis.", "Heme-granular or red blood cell casts", ">5 red blood cells/high power field. Exclude stone, infection, or other cause.", ">0.5 gram/24 hours", ">5 white blood cells/high power field. Exclude infection.", "Inflammatory type rash", "Abnormal, patchy or diffuse loss of hair", "Oral or nasal ulcerations", "Pleuritic chest pain with pleural rub or effusion or pleural thickening", "Pericardial pain with at least 1 of the following: rub, effusion, or electrocardiogram or echocardiogram confirmation", "Decrease in CH50, C3, or C4 below the lower limit of normal for testing laboratory", "Increased DNA binding by Farr assay above normal range for testing laboratory", ">38° C. Exclude infectious cause", "<100,000 platelets/× 10^9/L, exclude drug causes", "<3000 white blood cells/× 10^9/L, exclude drug causes."])
SLEDAI_weight = np.array([8, 8, 8, 8, 8, 8, 8, 8, 4, 4, 4, 4, 4, 4, 2, 2, 2, 2, 2, 2, 2, 1, 1, 1])
SLEDAI_ans = np.array(None, (object, [24]))
SLEDAI_ans_num = SLEDAI_ans.copy()

# SLEDAI_QA問答與答案紀錄完成  SLEDAI_cond, SLEDAI_def, SLEDAI_ans, SLEDAI_ans_num)
def SLEDAI_QA(CondArray, DefArray): 
  global SLEDAI_ans, SLEDAI_ans_num
  print(f'''--------------------------------------------------------------------------------------------------------------------
\nGiven that the patient probably suffers from SLE, we would like to further measure the disease activity baesd on "SLEDAI-2K(30 Days).2010".
Hopefully, we can help you understand the patient's current condition and even predict the outcome and prognosis.\n''')
  #問答前詳細說明判定標準
  print('''*Here are something you should be informed of before the questionaire. 
  1. Enter "Y" if descriptor in SLEDAI-2K sheet is present at the time of the visit or in the preceding 30 days.
  2. Please input correct key words for the App to run smoothly! (case-insensitive大小寫不拘)
  "Y" stands for "Yes"(i.e. a patient meets the condition); "N" for "No"(i.e. one doesn't meet the criteria); "NI" for "no information" 
  "B" or "F" respectively for "jump *Back or *Forward to another question"  
  3. case-insensitive大小寫不拘
  4. Detailed definition is showed below each question. Please refer to it if necessary.''')
  i = 0
  while i < len(CondArray):
      print(f'\n{i+1}. {CondArray[i]} ? (Current Ans: "{SLEDAI_ans[i]}")')     #是否附註該condition的定義給使用者作確認(v)/back&forward(V)/預設no information(v)
      print(f'(Definition: {DefArray[i]})')    #呈現出目前答題狀況(v)(*惟最後一題無法更改>考慮新增一項for最後檢查?)
      check = input(f'(Y/ N/ NI/ B/ F)  ').strip().upper()  #str.upper()是小寫轉大寫；str.capitalize()是字首大寫其餘小寫
      if check == "Y":
          SLEDAI_ans[i] = check   #先用Y、N組成ans_array，之後再process為ans_Number_array(0, 1)，再與SLEDAI_weight相乘，得到ans_Weight_array
          SLEDAI_ans_num[i] = 1
          i = i + 1
      elif check == "N":
          SLEDAI_ans[i] = check
          SLEDAI_ans_num[i] = 0
          i = i + 1
      elif check == "NI":  
          SLEDAI_ans[i] = check  ##這邊有改
          SLEDAI_ans_num[i] = 0
          i = i + 1 
      elif check == "B": 
          if i == 0:        #當輸入B或F時，讓使用者知道自己正跳至別題檢查，並可更正答案(v)(*希望呈現出目前答案給使用者參考?)
            print("*This is the first question!") 
          else:
            i -= 1       
      elif check == "F":
          if i == len(CondArray) - 1:
            print("*This is the last question!")  
          else:
            i += 1        
      else:
          print("Wrong key word!")
          i = i         
  return SLEDAI_ans, SLEDAI_ans_num

##SLEDAI_ans confirmation
def SLEDAI_Ans_Confirmation(CondArray):   
  global SLEDAI_ans, SLEDAI_ans_num
  # 1. 呈現出目前答題狀況(題號, condition, ans)
  print('\n* The following dataframe shows your current answers to each question, and we\'ll help you confirm them before the next step.')
  print('Q_Num  ', "Condition".ljust(70), 'Current_Ans\n', '-' * 125)
  for i in range(len(CondArray)):
    print(f'{i+1:<8}{CondArray[i]:<70} {SLEDAI_ans[i]}')
  # 2. 修正
  confirm2 = None
  while confirm2 != 'Y':
    confirm1 = None
    while confirm1 != 'Y' and confirm1 != 'N':
      confirm1 = input('Is there anything wrong that you\'d like to modify? (Please input "Y" or "N")').strip().upper()
    while confirm1 == 'Y':
      while True:
        QNum1 = input('Please input the question number in correct format! (eg. 1, 3, or 17 ; one at a time!)').strip()
        if ',' in QNum1:
          qnlist1 = []
          for qn in QNum1.split(','):
            qnlist1.append(qn.strip())
        else:
          qnlist1 = [QNum1]    
        try:
          for k in range(len(qnlist1)):
            qnlist1[k] = int(qnlist1[k]) - 1
          if 0 <= qnlist1[k] < len(CondArray):
            break
          else:
            print('Wrong question number!')
            continue  
        except Exception as ex:   # Exception為各種exceptions的superclass；將此exception存在ex變數中!!!  
          print('Wrong key word, please try again!') 
          print('Error_type:', type(ex))          
          print('Error!', ex)  
      for i in qnlist1:   
        print(f'({i+1}) {CondArray[i]} ? (Current Ans: "{SLEDAI_ans[i]}")')      
        while True:
          check = input(f'(Y/ N/ NI)  ').strip().upper()
          if check == "Y":
              SLEDAI_ans[i] = check   #先用Y、N組成ans_array，之後再process為ans_Number_array(0, 1)，再與SLEDAI_weight相乘，得到ans_Weight_array
              SLEDAI_ans_num[i] = 1
              break
          elif check == "N":
              SLEDAI_ans[i] = check
              SLEDAI_ans_num[i] = 0
              break
          elif check == "NI":
              SLEDAI_ans[i] = check   ##
              SLEDAI_ans_num[i] = 0
              break
          else:
              print("Wrong key word, please try again!")
      confirm1 = input('Is there anything else wrong that you\'d like to modify? (Please input "Y" or "N")').strip().upper()
    #再次確認目前ans 
    print('Again! The following dataframe shows your current answers to each question. Please double check them!') 
    print('Q_Num  ', "Condition".ljust(70), 'Current_Ans\n', '-' * 125)
    for i in range(len(CondArray)):
      print(f'{i+1:<8}{CondArray[i]:<70} {SLEDAI_ans[i]}')
    #最後確認
    confirm2 = None
    while confirm2 != 'Y' and confirm2 != 'N':
      confirm2 = input('Have you already confirmed that all answers correspond with your current health status? (Please input "Y" or "N")').strip().upper()
  print('Answer confirmation has finished! Please wait a moment for the results!')  
  return SLEDAI_ans, SLEDAI_ans_num

###處理各domains的weights(適用於clinical & immunological) 
## 先用Y、N組成ans_array(v)，問答時已產生SLEDAI_ans_num(0無症狀或無資料, 1有症狀)，將與SLEDAI_weight相乘，得到ans_Weight_array
# ans_Number_array與SLEDAI_weight相乘，得到SLEDAI_ans_Weight_array
def SLEDAI_Process():
    global SLEDAI_TrueWeight, SLEDAI_total
    SLEDAI_TrueWeight = np.multiply(SLEDAI_ans_num, SLEDAI_weight)  #SLEDAI_ans_num有問題
    SLEDAI_total = np.sum(SLEDAI_TrueWeight)  #計算總分
    print(f'SLEDAI_total_score = {SLEDAI_total}pts')
    return SLEDAI_TrueWeight, SLEDAI_total
#SLEDAI_Class(嚴重度分級)
def SLEDAI_Class(SLEDAI_Total):
  global SLEDAIClass
  if SLEDAI_Total == 0:
    SLEDAIClass = 'no activity'
  elif SLEDAI_Total < 6:
    SLEDAIClass = 'mild activity'
  elif SLEDAI_Total < 11:
    SLEDAIClass = 'moderate activity(Suggestion: greater than 50% probability of initiating therapy)'
  elif SLEDAI_Total < 20: 
    SLEDAIClass = 'high activity(Suggestion: greater than 50% probability of initiating therapy)' 
  else:
    SLEDAIClass = 'very high activity(Suggestion: greater than 50% probability of initiating therapy)' 
  print(f'*Judgement Result: {SLEDAIClass}')
  return SLEDAIClass
# No_Information_Suggestion2 
def NI_Suggestion2(Ans, Cond):  # 參數AnsArray（SLEDAI_ans）；ConditionArray(SLE_condition) 
  print(f'\n*The conditions with no information, if any, are as follows:') 
  NI_index = []
  NI_list = []
  for i in range(len(Ans)):
    if Ans[i] == "NI":
      NI_index.append(i)
  for i in NI_index:
    print(Cond[i])
    NI_list.append(Cond[i])    
  print(f'\nIf you have some questions with NO INFORMATION, those conditions are considered negative automatically by the App and thus it is likely that the patient\'s disease severity is underestimated. In this case, We suggest that you take further examinations to make up the dificiency and then you can get more accurate predictions from the App! Thanks!')
  return NI_list  
# Excel_automation
def ExcelAuto(): 
    global sickid, name, titer, status3
    from datetime import datetime
    t = datetime.now()
    while True:
        will = input("Would you like to save your data? (Y or N)").strip().upper()
        if will == "Y":
            anslist11 = [sickid, name, titer]
            if 0 < titer <= 80:                #將titer>80的病患資料一併記錄(在下面輸入時設計)
                for i in SLE_ans:
                    for j in i:
                        if j is None:
                            pass
                        else:
                            anslist11.append(j)                           
            anslist12 = [sickid, name]
            if 0 < titer <= 80:
                anslist12.append(total_score)
                anslist12.append(sleresult)
                anslist12.extend(Domain_weight)

            anslist21 = [sickid, name]
            if status3 == 'Y':  #Qbeforedai
                anslist21.extend(SLEDAI_ans)
                anslist21.append(t)
            anslist22 = [sickid, name]
            if status3 == 'Y':  #Qbeforedai
                anslist22.append(SLEDAI_total)
                anslist22.append(SLEDAIClass)
                anslist22.extend(SLEDAI_TrueWeight)
                anslist22.append(t)
            import openpyxl
            import os
            # os.chdir 是 python 切換到電腦指定路徑的方法
            os.chdir(r"E:\國防醫學院 醫學系\課業\lab\python project\總結")
            # 請填寫自己電腦裡Excel檔案的絕對路徑，填寫要處理的Excel檔案名稱
            wb = openpyxl.load_workbook('SLE_auto_data.xlsx')
            # sheet = wb.worksheets[0]
            sheet11 = wb['SLE_Ans']
            sheet12 = wb['SLE_Ans_weight']  
            sheet21 = wb['SLEDAI_Ans']
            sheet22 = wb['SLEDAI_Ans_weight']
            #使用for loop掃描所有A欄(姓名)，如果出現重複就更新舊資料，否則新增一個資料列
            check = 0
            for i in range(2, sheet11.max_row + 1):
                if sheet11.cell(i, 1).value == sickid:
                    for j in range(1, len(anslist11) + 1):
                        sheet11.cell(i, j).value = anslist11[j-1]
                    sheet11.cell(i, sheet11.max_column).value = t
                    for k in range(1, len(anslist12) + 1):    
                        sheet12.cell(i, k).value = anslist12[k-1]
                    sheet12.cell(i, sheet12.max_column).value = t
                    # if status3 = 'Y':        
                else:
                    check += 1
            if check == sheet11.max_row - 1:
                sheet11.append(anslist11)
                sheet12.append(anslist12)
                sheet11.cell(sheet11.max_row, sheet11.max_column).value = t
                sheet12.cell(sheet11.max_row, sheet12.max_column).value = t
            
            if status3 == 'Y':
                check1 = 0
                for i in range(2, sheet21.max_row + 1):
                    if sheet21.cell(i, 1).value == sickid:
                        for j in range(1, len(anslist21) + 1):
                            sheet21.cell(i, j).value = anslist21[j-1]
                        for k in range(1, len(anslist22) + 1):    
                            sheet22.cell(i, k).value = anslist22[k-1]
                    else:
                        check1 += 1 
                if check1 == sheet21.max_row - 1:
                    sheet21.append(anslist21)
                    sheet22.append(anslist22)              
            # 將結果存檔
            wb.save('SLE_auto_data.xlsx')
            break 
        elif will == "N":
            break
        else:
            print("Please input Y or N!")
            continue  